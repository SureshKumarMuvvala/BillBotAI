"""
excel_service.py — build styled openpyxl workbook from OCR results.
Extracted from notebook export-excel cell. All Colab/IPython code removed.
Returns an in-memory BytesIO; no disk I/O.
"""

import io
import logging
from pathlib import Path
from datetime import datetime
from typing import Any, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styles (preserved from notebook) ─────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="2D3E50")
ALT_FILL   = PatternFill("solid", fgColor="EBF5FB")
WHT_FILL   = PatternFill("solid", fgColor="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="EBF5FB")
OK_FILL    = PatternFill("solid", fgColor="D5F5E3")
ERR_FILL   = PatternFill("solid", fgColor="FADBD8")

HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=12, color="2D3E50")

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR    = Alignment(horizontal="center", vertical="center")
LFT    = Alignment(horizontal="left",   vertical="center")


def _cell(ws, row, col, value="", font=None, fill=None, align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = BORDER
    return c


def write_table(ws, headers, rows, start_row=1):
    for ci, h in enumerate(headers, 1):
        _cell(ws, start_row, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CTR)
    for ri, row_data in enumerate(rows, 1):
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        for ci, val in enumerate(row_data, 1):
            _cell(ws, start_row + ri, ci, val, font=BODY_FONT, fill=fill, align=LFT)
    for ci, h in enumerate(headers, 1):
        values = [str(h)] + [str(r[ci-1]) if ci-1 < len(r) else "" for r in rows]
        ws.column_dimensions[get_column_letter(ci)].width = max(
            min(max(len(v) for v in values) + 3, 40), 10
        )
    return start_row + len(rows)


def title_block(ws, text, subtitle="", merge_cols=11):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = LFT
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_cols)
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(name="Arial", italic=True, size=9, color="555555")
        c2.alignment = LFT


def build_workbook(
    job_id: str,
    filename: str,
    line_items: List[dict],
    quality_score: Any = None,
    latency_sec: float = 0.0,
) -> io.BytesIO:
    """Return an in-memory .xlsx stream. Mirrors notebook export-excel logic."""
    wb      = openpyxl.Workbook()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Sheet 1 — Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    title_block(ws_sum, "Chandra OCR — GRN Extraction Summary",
                subtitle=f"Generated: {now_str}")
    write_table(ws_sum, [
        "#", "File Name", "Convert", "Extract",
        "Quality Score", "Line Items", "Processing (s)", "Job ID",
    ], [[
        1, filename, "OK",
        "OK" if line_items else "FAIL",
        f"{quality_score}/5" if quality_score is not None else "—",
        len(line_items), f"{latency_sec:.1f}", job_id,
    ]], start_row=4)

    # Sheet 2 — All Line Items
    ws_all = wb.create_sheet("All Line Items")
    title_block(ws_all, "All GRN Line Items")
    all_headers = [
        "Source File", "Medicine Name", "Batch No", "Expiry Date",
        "Quantity", "Free Quantity", "Rate", "MRP", "GST %", "Amount", "HSN Code",
    ]
    all_rows = [[
        filename,
        item.get("medicine_name"), item.get("batch_no"), item.get("expiry_date"),
        item.get("quantity"), item.get("free_quantity"), item.get("rate"),
        item.get("mrp"), item.get("gst_percent"), item.get("amount"), item.get("hsn_code"),
    ] for item in line_items]
    if all_rows:
        write_table(ws_all, all_headers, all_rows, start_row=4)
    else:
        ws_all.cell(row=4, column=1).value = "No items extracted."

    # Sheet 3 — Per-file (mirrors notebook per-file section)
    stem = Path(filename).stem[:24]
    safe = "".join(c if c not in r'\/*?[]:"' else "_" for c in stem)
    ws   = wb.create_sheet(safe[:28])
    title_block(ws, f"{filename} — Line Items", subtitle=f"Rows extracted: {len(line_items)}")
    per_headers = ["Medicine Name","Batch No","Expiry Date","Quantity","Free Quantity",
                   "Rate","MRP","GST %","Amount","HSN Code"]
    per_rows = [[
        item.get("medicine_name"), item.get("batch_no"), item.get("expiry_date"),
        item.get("quantity"), item.get("free_quantity"), item.get("rate"),
        item.get("mrp"), item.get("gst_percent"), item.get("amount"), item.get("hsn_code"),
    ] for item in line_items]
    if per_rows:
        write_table(ws, per_headers, per_rows, start_row=4)
    else:
        ws.cell(row=4, column=1).value = "No items extracted."

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    logger.info("excel_service: workbook built — %d item(s)", len(line_items))
    return stream
